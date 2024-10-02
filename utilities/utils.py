import csv
import inspect
import logging

import softest
from openpyxl import Workbook, load_workbook


class Utils(softest.TestCase):
    def assertwithtext(self, list, value):
        # Assertion:- checking 1 stop text is present in all lists
        for values in list:
            self.soft_assert(self.assertIn, value, values.text, "Check whether vlue is present in assertion ")
            # assert value in values.text

        print("assert passed")
        print(len(list))
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        # Set class/method name from where it's called
        logger_name = inspect.stack()[1][3]

        # Create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        # Create a file handle amd set the log level and for console handler

        fh = logging.FileHandler("automation.log", mode="w")
        # Create formatter to show how my logs to be formatted
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - % (name)s-% (message)s',
                                      datefmt='%m/%d/%Y %I:%M:%S %p')

        # Add formatter to above console or file handler
        fh.setFormatter(formatter)
        # Add console or file handler to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    import csv
    def read_data_from_csv(filename):
        #create an emplty list
        datalist = []
        # open the file
        csvfile = open(filename,"r")
        # read the csv file
        reader = csv.reader(csvfile)
        #skip the headers from file
        next(reader)

        # Add values to datalist
        for rows in reader:
            datalist.append(rows)
        return datalist
